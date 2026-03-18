from collections import defaultdict
from typing import Dict, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

"""
Returns (wins, losses) for the tournament
"""
def tabulate_single_event(event_data):
    wins = 0
    losses = 0
    
    rounds = event_data['rounds']
    for round in rounds:
        if round['is_win']:
            wins += 1
        else:
            losses += 1

    return (wins, losses)


"""
Count every possible result in the player's history (e.g: player had 3 3-1s, etc.)
"""
def tabulate_results(events_data) -> None:
    player_results = defaultdict(int)
    player_results_by_losses = defaultdict(int)

    for event in events_data:
        result = tabulate_single_event(event)
        player_results[result] += 1
    sorted_results = dict(sorted(player_results.items(), key=lambda x: x[0][1]))

    for result, qty in sorted_results.items():
        # Skipping tournaments that stores asked us to apply just for the record
        if result[0] == 0 and result[1] == 0: continue

        print(f"{result[0]}-{result[1]}: {qty}")
        player_results_by_losses[result[1]] += qty

    print("=============================")
    for result, qty in player_results_by_losses.items():
        print(f"X-{result}: {qty}")


def tabulate_all_winrates(events_data):
    players_results = dict()

    for event in events_data:
        for round in event['rounds']:
            if 'membership_number' not in round['opponent_users'][0]:
                continue
            else:
                user_bandai_id = round['opponent_users'][0]['membership_number']
                if user_bandai_id not in players_results:
                    if round['is_win']:
                        players_results[user_bandai_id] = [1, 0]
                    else:
                        players_results[user_bandai_id] = [0, 1]
                else:
                    if round['is_win']:
                        players_results[user_bandai_id][0] += 1
                    else:
                        players_results[user_bandai_id][1] += 1

    return players_results


def load_bandai_username_id() -> Dict[str,str]:
    username_map = dict()

    try:
        with open("bandai_username_map.txt") as f:
            for line in f:
                parts = line.strip().split(":")
                if len(parts) < 2:
                    continue
                username, banda_id = parts[0], parts[1]
                username_map[banda_id] = username
    except FileNotFoundError:
        pass

    return username_map

"""
To be used with the result of `tabulate_all_winrates`
"""
def print_player_results(player_results) -> None:
    username_map = load_bandai_username_id()

    sorted_results = dict(sorted(player_results.items(), key=lambda x: x[1][0] + x[1][1], reverse=True))

    col_name = max((len(username_map.get(pid, pid)) for pid in sorted_results), default=10)
    col_name = max(col_name, len("Player"))

    header = f"{'Player':<{col_name}}  {'W':>4}  {'L':>4}  {'Total':>5}  {'Win%':>6}"
    print(header)
    print("-" * len(header))

    for player_id, res in sorted_results.items():
        tag = username_map.get(player_id, player_id)
        wins, losses = res
        total = wins + losses
        winrate = (wins / total * 100) if total > 0 else 0
        print(f"{tag:<{col_name}}  {wins:>4}  {losses:>4}  {total:>5}  {winrate:>5.1f}%")

"""
Count results against single player
"""
def results_vs_player(player_bandai_id: str, events_data) -> Tuple[int, int]:
    wins, losses = (0, 0)

    for event in events_data:
        for round in event['rounds']:
            if 'membership_number' not in round['opponent_users'][0] or round['opponent_users'][0]['membership_number'] != player_bandai_id:
                continue
            if round['is_win']:
                wins += 1
            else:
                losses += 1

    return (wins, losses)


"""
Export player results to an Excel file
"""
def export_player_results_xlsx(player_results, output_path: str = "results.xlsx") -> None:
    username_map = load_bandai_username_id()
    sorted_results = dict(sorted(player_results.items(), key=lambda x: x[1][0] + x[1][1], reverse=True))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    center = Alignment(horizontal="center")

    headers = ["Player", "W", "L", "Total", "Win%"]
    col_widths = [24, 6, 6, 8, 8]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    alt_fill = PatternFill(start_color="EAF0FB", end_color="EAF0FB", fill_type="solid")

    for row_idx, (player_id, res) in enumerate(sorted_results.items(), start=2):
        tag = username_map.get(player_id, player_id)
        wins, losses = res
        total = wins + losses
        winrate = round(wins / total * 100, 1) if total > 0 else 0.0

        row_fill = alt_fill if row_idx % 2 == 0 else None
        values = [tag, wins, losses, total, winrate]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col > 1:
                cell.alignment = center
            if row_fill:
                cell.fill = row_fill

    wb.save(output_path)
    print(f"Exported to {output_path}")